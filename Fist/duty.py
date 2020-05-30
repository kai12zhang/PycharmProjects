import openpyxl
import jpholiday
from calendar import Calendar
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment

cal = Calendar()

# ワークブックを新規作成する
book = openpyxl.Workbook()

# シートを取得し名前を変更する
now_year = 2020
now_month = 5
now_yearMonth = '2020年' + str(now_month) + '月'
excleName = now_yearMonth + '勤務表' + '.xlsx'
path = 'C:\\zkai\\pythonTest\\'
w_list = ['月曜日', '火曜日', '水曜日', '木曜日', '金曜日', '土曜日', '日曜日']
fill_weekend = openpyxl.styles.PatternFill(patternType='solid', fgColor='CCD1D1', bgColor='CCD1D1')
fill_holiday = openpyxl.styles.PatternFill(patternType='solid', fgColor='FEF5E7', bgColor='FEF5E7')
fill_head = openpyxl.styles.PatternFill(patternType='solid', fgColor='AED6F1', bgColor='AED6F1')

alignment = Alignment(horizontal='center', vertical='center')
font = Font(size=11, bold=True, color='000000')
border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
sheet = book.active
sheet.title = now_yearMonth

# ヘッダ設定
sheet.column_dimensions['B'].width = 15.0
sheet.column_dimensions['D'].width = 20.0

sheet["B3"].value = '日付'
sheet["B3"].fill = fill_head
sheet["B3"].border = border
sheet["B3"].font = font
sheet["B3"].alignment = alignment
sheet["C3"].value = '曜日'
sheet["C3"].fill = fill_head
sheet["C3"].border = border
sheet["C3"].font = font
sheet["C3"].alignment = alignment
sheet["D3"].value = '祝日'
sheet["D3"].fill = fill_head
sheet["D3"].border = border
sheet["D3"].font = font
sheet["D3"].alignment = alignment

r = 4
for date in cal.itermonthdates(now_year, now_month):
    if date.month == now_month:
        if date.weekday() == 5 or date.weekday() == 6:
            sheet.cell(row=r, column=2).fill = fill_weekend
            sheet.cell(row=r, column=3).fill = fill_weekend
            sheet.cell(row=r, column=4).fill = fill_weekend
        elif jpholiday.is_holiday_name(date) is not None:
            sheet.cell(row=r, column=2).fill = fill_holiday
            sheet.cell(row=r, column=3).fill = fill_holiday
            sheet.cell(row=r, column=4).fill = fill_holiday
        sheet.cell(row=r, column=2).value = date
        sheet.cell(row=r, column=2).border = border
        sheet.cell(row=r, column=3).value = w_list[date.weekday()]
        sheet.cell(row=r, column=3).border = border
        sheet.cell(row=r, column=4).value = jpholiday.is_holiday_name(date)
        sheet.cell(row=r, column=4).border = border
        r = r + 1

book.save(path + excleName)
